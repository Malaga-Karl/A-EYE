import syntax
import csv
import pandas as pd
import lexer
import openpyxl
from openpyxl.styles import PatternFill

# exception si 252
# if want niyo try, download https://docs.google.com/spreadsheets/d/1xzn_b8FHk7-cKC7gAeeUlDnoTKlbY8DXVyR3XfJoj6o/edit#gid=1770890317 as csv
# then rename it to aeyetest.csv and put it in A-EYE folder (bali marereplace niya ung current)
# run this and check results.xlsx for results
df = pd.read_csv('aeyetest.csv')

def analyze_syntax(code):
    result, errors = lexer.analyze_text(code)
    if not errors:
        syntax_result = syntax.analyze_syntax(result)

    if errors:
        for error in errors:
            return error.as_string()
    else:
        return syntax_result


string = df['test_item'][0]
with open('file.csv', 'w', newline='') as file:
    writer = csv.writer(file)
    writer.writerow(['Test Item', 'Result'])
    for index, row in df.iterrows():
        string = row['test_item']
        delims = ['\n', '\r', '\t']
        for delimiter in delims:
            string = " ".join(string.split(delimiter))
        result = analyze_syntax(string)
        if result != 'Syntax analysis successful':
            detail = 'Syntax Error'
        else:
            detail = 'No Syntax Error'
        
        writer.writerow([string, result])

file = pd.read_csv('file.csv', encoding='unicode_escape')
results = pd.ExcelWriter('results.xlsx')
file.to_excel(results, index=False)
results.save()
wb = openpyxl.load_workbook('results.xlsx')
ws = wb['Sheet1']

passed = PatternFill(patternType='solid', fgColor='32a862')
fail = PatternFill(patternType='solid', fgColor='ff0000')
for i in range(2, len(df) + 2):
    if ws[F'B{i}'].value == 'Syntax analysis successful':
        ws[F'B{i}'].fill = passed
    else:
        ws[F'B{i}'].fill = fail

wb.save('results.xlsx')

number_of_passed = file.value_counts('Remarks')['Pass']
print(f'Passing Rate: {round(number_of_passed/len(file)*100,2)}%')

